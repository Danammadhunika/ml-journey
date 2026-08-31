"""
tests/test_export_service.py
------------------------------
Smoke test for database/export_service.py -- confirms the tracker
actually gets written out as a real, readable .xlsx workbook with the
right columns and row data.
"""

from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from agents.job_matcher import LLMJobAnalysis, MatchResult
from database.export_service import export_tracker_to_excel
from database.models import Base
from database.tracker_service import log_company_update, save_match_result


@pytest.fixture()
def session():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
    db_session = SessionLocal()
    try:
        yield db_session
    finally:
        db_session.close()


def test_export_tracker_to_excel_writes_all_rows(session, tmp_path: Path):
    result = MatchResult(
        overall_score=90,
        technical_skills_score=90,
        experience_score=90,
        education_score=90,
        seniority_score=90,
        location_score=90,
        sponsorship_compatibility="COMPATIBLE",
        sponsorship_note=None,
        strengths=["Python"],
        gaps=[],
        missing_requirements=[],
        recommendation="APPLY",
        reason="Great fit.",
        llm_analysis=LLMJobAnalysis(summary_reason="Great fit."),
    )
    save_match_result(
        session, job_title="Python Developer", job_description_text="JD", match_result=result, company="Acme"
    )
    log_company_update(session, "HCL", status="RECRUITER_SCREENING", notes="Screening done.")
    session.commit()

    output_path = tmp_path / "tracker.xlsx"
    export_tracker_to_excel(session, output_path)

    assert output_path.exists()
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook["Applications"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert "Company" in header
    assert "Current Status" in header
    assert "Follow-up Date" in header

    companies = {row[header.index("Company")] for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert companies == {"Acme", "HCL"}
